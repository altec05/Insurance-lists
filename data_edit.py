import os
from openpyxl.utils.cell import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
import win32com.client as win32
import time
import psutil
from tkinter import messagebox

import path_funcs
import service
import datetime

# import variables
import variables as vars
import messages as mes
import check_funcs as chf
# from path_funcs import get_output_path


# Полная очистка всех глобальных переменных для последующей обработки
def clear_vars_full():
    vars.tab_hospital_path = ''
    vars.tab_visit_path = ''
    vars.tab_donors_path = ''
    vars.out_tab_path = ''
    vars.temp_hospital_tab_path = ''
    vars.temp_visit_tab_path = ''
    vars.donors_not_uniq_inp = 0
    vars.donors_not_uniq_outp = 0
    vars.donors_get = 0
    vars.donors_with_multiple_dons = 0
    vars.donors_out = 0
    vars.donors_with_dots = 0
    vars.donors_nums_indexes_list.clear()
    vars.tab_get_time_min = 0
    vars.tab_get_time_sec = 0
    vars.tab_fill_time = 0
    vars.tab_data_row_last = ''
    vars.tab_data_row_hospital = ''
    vars.tab_data_row_visit = ''
    vars.data_for_name = ''
    vars.first_row_text = ''
    vars.data_lose = False
    vars.donors_size_lose = False
    vars.donors_full_fio_lose = False
    vars.bad_datas.clear()
    vars.bad_sizes.clear()
    vars.bad_donors_fio.clear()
    vars.log_all_path = ''
    vars.city_name = ''


# Обработка входных данных и заполнений итоговой таблицы
def get_input_data():
    # Замеряем время обработки таблиц
    t_start = time.perf_counter()

    # Читаем данные из входящих таблиц и заносим в переменные,
    # Получаем путь до временного файла с расширением xlsx
    # hospital_data = get_donors_list(vars.tab_hospital_path, 'hospital')
    hospital_data = get_donors_list(vars.tab_hospital_path, 'hospital')

    if vars.tab_visit_path == '':
        visit_data = ''
    else:
        visit_data = get_donors_list(vars.tab_visit_path, 'visit')

    donors_data = get_donors_list(vars.tab_donors_path, 'donors')

    # Если одна из таблиц не обработалась, то завершаем
    # Ошибка выдается внутри обработки таблицы
    if hospital_data == False or visit_data == False or donors_data == False:
        clear_vars_full()
        vars.end_func = True
        return False

    # Итоговое время чтения входных данных
    all_time = time.perf_counter() - t_start
    sec = round(all_time)
    delta = delta = str(datetime.timedelta(seconds=sec))
    vars.tab_get_time_sec = delta

    # Замеряем начало заполнения итоговой таблицы и её форматирования
    t_start = time.perf_counter()

    # Заполняем и изменяем итоговую таблицу
    fill_donors_tab(vars.tab_donors_path_new, hospital_data, visit_data, donors_data)

    # Итоговое время заполнения и форматирования выходной таблицы
    all_time = time.perf_counter() - t_start
    sec = round(all_time)
    delta = str(datetime.timedelta(seconds=sec))
    vars.tab_fill_time_sec = delta

    statistic_data = vars.get_statistic()

    mes.info('Завершение заполнения обработки данных',
             f'Статистика:\n\n1. Время обработки таблиц: {statistic_data[0]}\n\n2. Время заполнения полученных данных и форматирования итоговой таблицы: {statistic_data[1]}\n\n3. Записей обработано: {statistic_data[2]} шт.\n4. Из них с несколькими донациями: {statistic_data[3]} шт.\n5. Итого записей получено: {statistic_data[4]} шт.\n\nСейчас откроется папка с итоговым файлом.')

    # По завершению открываем папку с полученными файлами
    os.system(
        fr"explorer.exe {vars.tab_donors_path_new.replace(os.path.basename(vars.tab_donors_path_new), '')}")

    clear_vars_full()
    vars.end_func = True
    return True


# Обработка входной таблицы
def get_donors_list(path, name):
    main_path = ''
    # Получаем расширение таблицы
    file_extension = chf.get_file_info(path)[2]
    # Проверяем, что xls/xlsx
    if chf.check_extension(file_extension):
        # Повторно проверяем запущенный Excel и закрываем
        if chf.check_run_excel():
            while chf.check_run_excel():
                for proc in psutil.process_iter():
                    if proc.name() == 'EXCEL.EXE':
                        try:
                            proc.kill()
                        except:
                            print(f'Пропустил процесс: {proc.name()}')
                            continue

        # Пересохраняем в xlsx с нужным именем и запоминаем новые пути
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(path)
        wb.DisplayAlerts = False
        if name == 'donors':
            vars.tab_donors_path_new = path_funcs.get_out_file_path(path)
            main_path = vars.tab_donors_path_new
        else:
            if name == 'hospital':
                vars.tab_hospital_path_new = path_funcs.get_temp_file_path(path)
                main_path = vars.tab_hospital_path_new
            elif name == 'visit':
                vars.tab_visit_path_new = path_funcs.get_temp_file_path(path)
                main_path = vars.tab_visit_path_new
        wb.SaveAs(main_path, FileFormat=51)
        wb.Close()
        excel.Application.Quit()

        # Открываем таблицу по полученному пути
        wb = load_workbook(main_path)

        # Получаем активный лист
        ws = wb.active

        # Для записи строк таблицы
        list_temp = []

        # Для записи 1-4 строк по 4 колонку таблицы в список для получения заголовка
        list_data_temp = []
        list_data_temp = list(ws.iter_rows(min_row=1, max_row=4, max_col=4))

        # Для перезаписи списка с удалением None строк
        clear_list_temp = []
        for row in list_data_temp:
            for col in row:
                if col.value is not None:
                    clear_list_temp.append(col.value)

        # Записываем в строку для дальнейшего поиска в титуле
        first_row_text = str(clear_list_temp)

        # Получаем название города в таблице
        city_name = service.get_city_name_from_row(str(clear_list_temp[0]))
        if city_name != '':
            vars.city_name = city_name
        else:
            vars.city_name = 'Красноярск'

        # Получаем индексы периода выгрузки в титуле таблицы
        start_index = first_row_text.index('за период с') + 10
        end_index = first_row_text.index('по') + 13

        # Заполняем периоды выгрузки для таблиц
        if name == 'hospital':
            vars.tab_data_row_hospital = first_row_text[start_index:end_index]
            vars.data_for_name = first_row_text[first_row_text.index('по') + 3:end_index]
        elif name == 'visit':
            vars.tab_data_row_visit = first_row_text[start_index:end_index]
            vars.data_for_name = first_row_text[first_row_text.index('по') + 3:end_index]
        elif name == 'donors':
            vars.first_row_text = str(clear_list_temp[0])
            vars.tab_data_row_last = first_row_text[start_index:end_index]
            vars.data_for_name = first_row_text[first_row_text.index('по') + 3:end_index]

        # Записываем данные из таблиц построчно
        if name == 'hospital' or name == 'visit':
            list_temp = list(ws.iter_rows(min_row=7, max_col=13, max_row=ws.max_row + 1))
        elif name == 'donors':
            list_temp = list(ws.iter_rows(min_row=1, max_col=20, max_row=ws.max_row + 1))

        # Для записи данных из таблиц из списка в строку одним списком
        list_of_lists = list(map(list, list_temp))

        # Для очистки от None строк списков
        clear_list_of_lists = []

        if name == 'hospital' or name == 'visit':
            for row in list_of_lists:
                if str(row[0].value).isdigit():
                    clear_list_of_lists.append(row)
        else:
            for row in list_of_lists:
                clear_list_of_lists.append(row)

        wb.close()

        # Переписываем построчно каждую ячейку в итоговый список
        clear_data_list = []
        for row in clear_list_of_lists:
            temp_list = []
            for col in row:
                try:
                    temp_list.append(col.value)
                except:
                    temp_list.append(col)
            clear_data_list.append(temp_list)

        if name == 'hospital' or name == 'visit':
            service.del_file(main_path)

        return clear_data_list
    else:
        mes.error('Проверка формата входных данных', 'Внимание!'
                                                     '\n\nУказанные таблицы должны иметь расширение "xlsx" или "xls"!')
        return False


# Заполняем и форматируем итоговую таблицу
def fill_donors_tab(path, hospital_list, visit_list, donors_list):
    # Число доноров в УП-3
    counter_donors = 0
    # Число измененных донаций
    counter_donors_edit_dons = 0

    # Проверяем периоды выгрузки входных таблиц
    chf.check_upload_period()

    # Для хранения и разбора полученных списков из таблиц
    inp_data_list = []

    if visit_list != '':
        inp_data_list = [hospital_list, visit_list]
    else:
        inp_data_list = [hospital_list]

    # Удаляем доп. столбцы из-за кривых входных данных со скрытыми колонками
    for row in donors_list:
        del row[0]

    for row in donors_list:
        del row[4]
        del row[17]
        del row[16]
        del row[13]
        del row[12]
        del row[10]
        del row[8]

    # Подсчет количества записей в УП-3
    for row in donors_list:
        if not str(row[0]).isdigit():
            continue
        else:
            counter_donors += 1


    # # Списки для записи одинаковых, но разных по номеру
    # namesakes_list_inp = list()
    # namesakes_list_outp = list()

    # Списки для записи одинаковых по д.номеру и фио, но разных по году
    number_name_bad_inp = list()
    number_name_bad_outp = list()

    # Список для записи доноров с несколькими донациями
    multiple_donors_list = list()

    # Разбираем каждый из входящих УП2 файлов
    for inp_list in inp_data_list:
        # разбираем их построчно
        for row in inp_list:
            # Если строка без № п/п, то пропускаем
            if not str(row[0]).isdigit():
                print(f'Пропустил не число {str(row[0])}')
                continue
            else:
                # Получаем Фамилия И.О. из "Фамилия Имя Отчество" каждого из входящих файлов для поиска
                fio = service.get_fio(str(row[2]))

                # Ищем вхождения Ф И.О. в списке доноров
                indices = [(i, x.index(fio)) for i, x in enumerate(donors_list) if fio in x]
                if indices:
                    # берем год рождения из входящего списка
                    year = str(row[4])
                    for index in indices:
                        # Если одинаковые по донорскому номеру
                        if str(row[1]) == str(donors_list[index[0]][1]):
                            # Получаем год проверяемого из УП-3
                            finded_year = str(donors_list[index[0]][3].split('.')[2])
                            # Если одинаковые по году и ФИ.О.
                            if year == finded_year:
                                # заменили ФИО на полное
                                donors_list[index[0]][index[1]] = str(row[2])

                                # если донаций больше 1 и их две
                                if int(donors_list[index[0]][10]) == 2:
                                    # Записываем в список доноров с донациями больше 1
                                    temp_multiple_donor = [donors_list[index[0]][0], donors_list[index[0]][1],
                                                           donors_list[index[0]][2], donors_list[index[0]][3],
                                                           donors_list[index[0]][10], donors_list[index[0]][11]]
                                    multiple_donors_list.append(temp_multiple_donor)
                                    # заменяем кол-во на 1
                                    donors_list[index[0]][10] = donors_list[index[0]][10] - 1
                                    counter_donors_edit_dons += 1
                                    # дублируем строку и редактируем её
                                    temp_list = []
                                    temp_list = donors_list[index[0]].copy()

                                    # Проверка кол-ва донаций и их уравнивание
                                    # Если 1 и 2
                                    if donors_list[index[0]][7] > 0 and donors_list[index[0]][8] > 0:
                                        donors_list[index[0]][8] = donors_list[index[0]][8] - 1
                                        temp_list[7] = temp_list[7] - 1
                                    # Если 1 и 3
                                    elif donors_list[index[0]][7] > 0 and donors_list[index[0]][9] > 0:
                                        donors_list[index[0]][9] = donors_list[index[0]][9] - 1
                                        temp_list[7] = temp_list[7] - 1
                                    # Если 2 и 3
                                    elif donors_list[index[0]][8] > 0 and donors_list[index[0]][9] > 0:
                                        donors_list[index[0]][9] = donors_list[index[0]][9] - 1
                                        temp_list[8] = temp_list[8] - 1
                                    # Если 3 и 3
                                    elif donors_list[index[0]][9] > 1:
                                        donors_list[index[0]][9] = donors_list[index[0]][9] - 1
                                        temp_list[9] = temp_list[9] - 1
                                    # Если 1 и 1
                                    elif donors_list[index[0]][7] > 1:
                                        donors_list[index[0]][7] = donors_list[index[0]][7] - 1
                                        temp_list[7] = temp_list[7] - 1
                                    # Если 2 и 2
                                    elif donors_list[index[0]][8] > 1:
                                        donors_list[index[0]][8] = donors_list[index[0]][8] - 1
                                        temp_list[8] = temp_list[8] - 1
                                    donors_list.insert(index[0] + 1, temp_list)
                                # если донаций больше 1 и их три
                                elif int(donors_list[index[0]][10]) == 3:
                                    # Записываем в список доноров с донациями больше 1
                                    temp_multiple_donor = [donors_list[index[0]][0], donors_list[index[0]][1],
                                                           donors_list[index[0]][2], donors_list[index[0]][3],
                                                           donors_list[index[0]][10], donors_list[index[0]][11]]
                                    multiple_donors_list.append(temp_multiple_donor)
                                    # заменяем кол-во на 1
                                    donors_list[index[0]][10] = donors_list[index[0]][10] - 2
                                    counter_donors_edit_dons += 1
                                    # дублируем строку и редактируем её
                                    temp_list = []
                                    temp_list.append(donors_list[index[0]].copy())
                                    temp_list.append(donors_list[index[0]].copy())

                                    # Проверка кол-ва донаций и их уравнивание
                                    # Если 1 1 1
                                    if donors_list[index[0]][7] == 1 and donors_list[index[0]][8] == 1 and \
                                            donors_list[index[0]][9] == 1:
                                        donors_list[index[0]][8] = donors_list[index[0]][8] - 1
                                        donors_list[index[0]][9] = donors_list[index[0]][9] - 1
                                        temp_list[0][7] = temp_list[0][7] - 1
                                        temp_list[0][9] = temp_list[0][9] - 1
                                        temp_list[1][7] = temp_list[1][7] - 1
                                        temp_list[1][8] = temp_list[1][8] - 1
                                    # Если 2 1 0
                                    elif donors_list[index[0]][7] == 2 and donors_list[index[0]][8] == 1 and \
                                            donors_list[index[0]][9] == 0:
                                        donors_list[index[0]][7] = donors_list[index[0]][7] - 1
                                        donors_list[index[0]][8] = donors_list[index[0]][8] - 1
                                        temp_list[0][7] = temp_list[0][7] - 1
                                        temp_list[0][8] = temp_list[0][8] - 1
                                        temp_list[1][7] = temp_list[1][7] - 2
                                    # Если 2 0 1
                                    elif donors_list[index[0]][7] == 2 and donors_list[index[0]][8] == 0 and \
                                            donors_list[index[0]][9] == 1:
                                        donors_list[index[0]][7] = donors_list[index[0]][7] - 1
                                        donors_list[index[0]][9] = donors_list[index[0]][9] - 1
                                        temp_list[0][7] = temp_list[0][7] - 1
                                        temp_list[0][9] = temp_list[0][9] - 1
                                        temp_list[1][7] = temp_list[1][7] - 2
                                    # Если 1 2 0
                                    elif donors_list[index[0]][7] == 1 and donors_list[index[0]][8] == 2 and \
                                            donors_list[index[0]][9] == 0:
                                        donors_list[index[0]][8] = donors_list[index[0]][8] - 2
                                        temp_list[0][7] = temp_list[0][7] - 1
                                        temp_list[0][8] = temp_list[0][8] - 1
                                        temp_list[1][7] = temp_list[1][7] - 1
                                        temp_list[1][8] = temp_list[1][8] - 1
                                    # Если 1 0 2
                                    elif donors_list[index[0]][7] == 1 and donors_list[index[0]][8] == 0 and \
                                            donors_list[index[0]][9] == 2:
                                        donors_list[index[0]][9] = donors_list[index[0]][9] - 2
                                        temp_list[0][7] = temp_list[0][7] - 1
                                        temp_list[0][9] = temp_list[0][9] - 1
                                        temp_list[1][7] = temp_list[1][7] - 1
                                        temp_list[1][9] = temp_list[1][9] - 1
                                    # Если 3 0 0
                                    elif donors_list[index[0]][7] == 3 and donors_list[index[0]][8] == 0 and \
                                            donors_list[index[0]][9] == 0:
                                        donors_list[index[0]][7] = donors_list[index[0]][7] - 2
                                        temp_list[0][7] = temp_list[0][7] - 2
                                        temp_list[1][7] = temp_list[1][7] - 2
                                    # Если 0 3 0
                                    elif donors_list[index[0]][7] == 0 and donors_list[index[0]][8] == 3 and \
                                            donors_list[index[0]][9] == 0:
                                        donors_list[index[0]][8] = donors_list[index[0]][8] - 2
                                        temp_list[0][8] = temp_list[0][8] - 2
                                        temp_list[1][8] = temp_list[1][8] - 2
                                    # Если 0 0 3
                                    elif donors_list[index[0]][7] == 0 and donors_list[index[0]][8] == 0 and \
                                            donors_list[index[0]][9] == 3:
                                        donors_list[index[0]][9] = donors_list[index[0]][9] - 2
                                        temp_list[0][9] = temp_list[0][9] - 2
                                        temp_list[1][9] = temp_list[1][9] - 2

                                    donors_list.insert(index[0] + 1, temp_list[0])
                                    donors_list.insert(index[0] + 2, temp_list[1])
                            else:
                                print(f'Не одинаковые по году, но одинаковые по д.номеру и ФИО: {row[2]} {year} и {donors_list[index[0]][2]} {finded_year}')
                                number_name_bad_inp.append(row.copy())
                                number_name_bad_outp.append(donors_list[index[0]].copy())
                        else:
                            print('Не одинаковые по донорскому номеру')
                            # namesakes_list_inp.append(row.copy())
                            # namesakes_list_outp.append(donors_list[index[0]].copy())
                            continue
                else:
                    continue

    don_number = 1  # № п/п
    counter = 1

    # Заполняем шапку и задаем порядковые номера
    for row in donors_list:
        # Задаем порядковый номер записям
        if str(row[0]) is None:
            counter += 1
            continue
        elif counter == 4:
            row[0] = '№\nп/п'
            row[1] = 'Глоб. №'
            row[2] = 'ФИО\nдонора'
            row[3] = 'Дата\nрожд.'
            row[4] = 'Гр.\nкр.'
            row[5] = 'Ре-\nзус'
            row[6] = 'Перв.'
            row[7] = 'Донаций'
            row[11] = 'Место\nдонаций'
            counter += 1
        elif counter > 4:
            if counter == 5:
                row[7] = 'Кр'
                row[8] = 'п/ф'
                row[9] = 'Тр/ф'
                row[10] = 'Всего'
                counter += 1
                continue
            elif str(row[0]).isdigit():
                row[0] = don_number
                don_number += 1
                counter += 1
            else:
                counter += 1
                continue
        else:
            counter += 1
            continue

    # Сравнение размеров списков
    sizes = []
    for tab in inp_data_list:
        temp_size = 0
        for row in tab:
            if str(row[0]).isdigit():
                temp_size = row[0]
        sizes.append(temp_size)

    # Получаем общий размер записей из УП-2
    temp_size_all = 0
    for size in sizes:
        temp_size_all += size

    # Доноров обработано
    vars.donors_get = counter_donors
    vars.donors_with_multiple_dons = counter_donors_edit_dons

    # Сверяем количество записей в УП-2 и обработанных в УП-3
    if temp_size_all != counter_donors:
        print('Не совпадает ', temp_size_all, ' и ', counter_donors)
        vars.donors_size_lose = True
        vars.bad_sizes.append(('!!! Не совпадает число обработанных и общее число входных записей\n', '\nОбработанных в у-П3: ', counter_donors, '\nОбщее число входных записей у-П2: ', temp_size_all))
        mes.warning('Проверка количества записей в таблицах', f'Внимание!\n\nСумма доноров не сходится во входящих таблицах!\n\n'
                                                              f'Обработанных в у-П3: {counter_donors}'
                                                              f'\nОбщее число входных записей у-П2: {temp_size_all}')

    # Очищаем список индексов перед заполнением
    vars.donors_nums_indexes_list.clear()

    # Записываем в переменные для лога необработанные записи
    for row in donors_list:
        if row[2] is not None:
            if str(row[0]).isdigit():
                if str(row[2]).count('.') > 1:
                # if '.' in str(row[2]):
                    vars.donors_with_dots += 1
                    vars.donors_full_fio_lose = True
                    vars.donors_nums_indexes_list.append(str(row[0]))
                    vars.bad_donors_fio.append(('\n№ п/п ', str(row[0]), ' ФИО: ', row[2]))
        else:
            continue

    # Сообщаем о необработанных записях (сокращенное ФИО)
    if vars.donors_full_fio_lose:
        mes.warning('Проверка обработанных записей',
                    'Внимание!\n\nОбнаружены записи с сокращенным ФИО в итоговом файле!\nПроверьте файл отчета об ошибках или дождитесь предупреждения!')

    # Готовим список необработанных записей для сообщения пользователю
    if vars.donors_with_dots != 0:
        message = ''
        for indx in vars.donors_nums_indexes_list:
            message += indx + ' '
        mes.warning('Проверка обработанных записей', f'Внимание!\n\nОбнаружены записи в итоговом файле, которые не были обработаны!\n\nКоличество таких записей - {vars.donors_with_dots}.\n\nИндексы таких записей в итоговом файле:\n{message}.')

    # Записей получено
    vars.donors_out = don_number - 1

    # Вставляем в список итоговых данных дополнительные ячейки
    donors_list[1][10] = 'Форма 410/у-П3'
    donors_list[len(donors_list) - 1][4] = 'Зав. донорским отделением'
    donors_list[len(donors_list) - 3][6] = 'Итого:'

    # Считаем количество донаций по категориям и их итоговое количество
    blood_ct = 0
    plazma_ct = 0
    tromb_ct = 0
    all_ct = 0
    for row in donors_list:
        if row[7] is not None:
            if str(row[7]).isdigit():
                blood_ct += int(row[7])
                plazma_ct += int(row[8])
                tromb_ct += int(row[9])
                all_ct += int(row[10])

    # Заполняем итоговое количество донаций
    donors_list[len(donors_list) - 3][7] = blood_ct
    donors_list[len(donors_list) - 3][8] = plazma_ct
    donors_list[len(donors_list) - 3][9] = tromb_ct
    donors_list[len(donors_list) - 3][10] = all_ct

    # Переписываем отформатированные данные таблицы в новый список
    donors_data_list = []
    for row in donors_list:
        temp_list = []

        for col in row:
            temp_list.append(col)
        donors_data_list.append(temp_list)

    # Удаляем пустые строки
    del donors_data_list[5]
    del donors_data_list[5]

    # Создаем новую книгу
    try:
        wb = load_workbook(path)
    except Exception as e:
        mes.error('Ошибка начала обработки',
                  f'Внимание!\n\nПри обработке конвертируемых файлов произошла ошибка:\n\n{e}')

    # Удаляем пустой лист
    wb.remove(wb[wb.sheetnames[0]])

    # Создаем новый лист Итог
    wb.create_sheet('Итог', 0)
    ws = wb['Итог']

    # Заполняем лист отформатированными данными
    for row in donors_data_list:
        ws.append(row)

    # Вставляем строку заголовка
    ws.cell(row=1, column=2, value=vars.first_row_text)

    # Объединяем ячейки шапки таблицы
    merge_cells_list = (
        'B1:J3', 'K2:L2', 'A4:A5', 'B4:B5', 'C4:C5', 'D4:D5', 'E4:E5', 'F4:F5', 'G4:G5', 'H4:K4', 'L4:L5')
    for cells in merge_cells_list:
        ws.merge_cells(cells)

    # Определяем координаты строк источника донаций и объединяем каждую по длине
    list_temp = list(ws.iter_rows(min_row=0, max_col=2, max_row=ws.max_row + 1))
    coordinates = []
    counter = 1
    for row in list_temp:
        if str(row[0].value) is not None and str(row[0].value) != '' and not str(row[0].value).isdigit() and counter > 5 and counter < ws.max_row - 4:
            coordinates.append((row[0].coordinate, row[0].row))
            counter += 1
        else:
            counter += 1
            continue

    # Координаты в обратном порядке для безопасного удаления
    rev_coordinates = coordinates[::-1]

    # Удаляем строки источника донаций по координатам
    for coord in rev_coordinates:
        print(f'Удаляю {ws[coord[0]].value}: {coord[1]} = {coord[1]}')
        ws.delete_rows(coord[1])
        print('Удалил!')

    # Рисуем границы для ячеек и выравниваем текст
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row in range(ws.min_row, ws.max_row - 2):
        for col in range(ws.min_column, ws.max_column + 1):
            if row == 1 or row == 2 or row == 3:
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center",
                                                                   wrap_text=True)
            elif row == 4 or row == 5:
                ws.cell(row=row, column=col).border = thin_border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center",
                                                                   wrap_text=True)
            else:
                ws.cell(row=row, column=col).border = thin_border

    ws.column_dimensions[get_column_letter(1)].width = 6
    ws.column_dimensions[get_column_letter(2)].width = 12
    ws.column_dimensions[get_column_letter(3)].width = 30
    ws.column_dimensions[get_column_letter(4)].width = 10
    ws.column_dimensions[get_column_letter(5)].width = 5
    ws.column_dimensions[get_column_letter(6)].width = 5
    ws.column_dimensions[get_column_letter(8)].width = 5
    ws.column_dimensions[get_column_letter(9)].width = 5
    ws.column_dimensions[get_column_letter(10)].width = 5
    ws.row_dimensions[1].height = 42
    ws.row_dimensions[2].height = 42
    ws.row_dimensions[3].height = 42

    wb.save(path)

    dn_donors_inp, dn_donors_out = '', ''
    dn_inp_size, dn_out_size = 0, 0
    year_donors_inp, year_donors_out = '', ''
    year_inp_size, year_out_size = 0, 0
    multiple_donors_text = ''

    # # Если попались доноры, похожие по ФИО, но разные по д.номеру, то записываем их в файл
    # if len(namesakes_list_inp) > 0 and len(namesakes_list_outp) > 0:
    #     dn_donors_inp, dn_donors_out = service.get_not_uniq_donor_numbers(namesakes_list_inp, namesakes_list_outp)

    # Если попались доноры, похожие по ФИО и д.номеру, но разные по году, то записываем их в файл
    if len(number_name_bad_inp) > 0 and len(number_name_bad_outp) > 0:
        year_donors_inp, year_donors_out = service.get_not_uniq_donor_numbers(number_name_bad_inp, number_name_bad_outp)

    # Если доноры с несколькими донациями
    if vars.donors_with_multiple_dons > 0:
        multiple_donors_text = service.get_multiple_donors_text(multiple_donors_list)

    # Получаем итоговый текст для записи в файл
    out_text = service.get_full_text(dn_donors_inp, dn_donors_out, dn_inp_size, dn_out_size, year_donors_inp, year_donors_out, year_inp_size, year_out_size, multiple_donors_text)

    # Записываем или удаляем файл отчета
    if vars.multidons_out == True:
        service.write_file(out_text)

    # Записываем log файлы в папку и на сервер в общий log
    if vars.log_out == True:
        service.get_log_file()